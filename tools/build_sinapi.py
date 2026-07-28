#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Extrai os preços do Espírito Santo da planilha SINAPI_Referência e gera data/base-sinapi-es.json.

A planilha oficial traz as 27 UFs lado a lado:
  ISD (insumos)     — cabeçalho na linha 10; Classificação | Código | Descrição | Unidade | Origem | <1 coluna por UF>
  CSD (composições) — cabeçalho na linha 10; Grupo | Código | Descrição | Unidade | <par (Custo, %AS) por UF>

Usa as abas SEM desoneração (encargos sociais sobre a mão de obra), que é a base comparável
com a DER-ES: ambas são custo direto, BDI 0.

Uso:  python3 tools/build_sinapi.py <SINAPI_Referencia.xlsx> [--uf ES] [--out data/base-sinapi-es.json]
"""
import json, re, sys
import openpyxl

UFS = ['AC', 'AL', 'AM', 'AP', 'BA', 'CE', 'DF', 'ES', 'GO', 'MA', 'MG', 'MS', 'MT',
       'PA', 'PB', 'PE', 'PI', 'PR', 'RJ', 'RN', 'RO', 'RR', 'RS', 'SC', 'SE', 'SP', 'TO']
HEADER_ROW = 10

argv = [a for a in sys.argv[1:] if not a.startswith('--')]
uf = sys.argv[sys.argv.index('--uf') + 1].upper() if '--uf' in sys.argv else 'ES'
out = sys.argv[sys.argv.index('--out') + 1] if '--out' in sys.argv else 'data/base-sinapi-es.json'
src = argv[0]
if uf not in UFS:
    sys.exit(f"UF desconhecida: {uf}")
iuf = UFS.index(uf)

# data_only=False de propósito: na aba CSD o código da composição está dentro de uma fórmula
# =HYPERLINK(...;<código>) cujo resultado o Excel não gravou em cache (viria 0). Os preços são
# literais numéricos, então vêm corretos nos dois modos.
wb = openpyxl.load_workbook(src, read_only=True, data_only=False)
RE_COD = re.compile(r'MATCH\(\s*(\d+)')


def cabecalho(ws):
    """Devolve a linha de cabeçalho como lista, para conferir o alinhamento das colunas."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=HEADER_ROW, values_only=True), 1):
        if i == HEADER_ROW:
            return list(row)
    return []


def num(v):
    """Preço; devolve None quando em branco ou zero (SINAPI usa vazio/0 para 'não determinado')."""
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return float(v) or None
    try:
        return float(str(v).replace('.', '').replace(',', '.')) or None
    except ValueError:
        return None


def codigo(v):
    """Código da composição: literal, ou extraído da fórmula =HYPERLINK(...MATCH(<cod>...)...)."""
    if v is None:
        return ''
    if isinstance(v, str) and v.startswith('='):
        m = RE_COD.search(v)
        return m.group(1) if m else ''
    return str(v).strip()


# ── INSUMOS (ISD): 1 coluna por UF a partir da 6ª ──────────────────────────────
ws = wb['ISD']
col_ins = 5 + iuf                     # índice 0-based: col 6 = AC → 5 + posição da UF
hdr = cabecalho(ws)
assert str(hdr[col_ins]).strip().upper() == uf, f"ISD: coluna {col_ins} é {hdr[col_ins]!r}, esperado {uf}"
insumos = []
for i, row in enumerate(ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True), HEADER_ROW + 1):
    cod, desc, und = row[1], row[2], row[3]
    if not cod or not desc:
        continue
    p = num(row[col_ins])
    if p is None:                     # preço em branco = não determinado nesta UF
        continue
    insumos.append({"c": codigo(cod), "d": str(desc).strip(),
                    "u": str(und or '').strip(), "p": round(p, 2),
                    "cls": str(row[0] or '').strip()})

# ── COMPOSIÇÕES (CSD): par (Custo, %AS) por UF a partir da 5ª coluna ───────────
ws = wb['CSD']
col_comp = 4 + iuf * 2
hdr = cabecalho(ws)
assert 'Custo' in str(hdr[col_comp]), f"CSD: coluna {col_comp} é {hdr[col_comp]!r}, esperado Custo"
comps = []
for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
    cod, desc, und = row[1], row[2], row[3]
    if not desc:
        continue
    p = num(row[col_comp])
    if p is None:
        continue
    comps.append({"c": codigo(cod), "d": str(desc).strip(),
                  "u": str(und or '').strip(), "p": round(p, 2),
                  "grp": str(row[0] or '').strip(),
                  "as": num(row[col_comp + 1]) or 0})
wb.close()

base = {"meta": {"fonte": "SINAPI — Caixa Econômica Federal", "uf": uf,
                 "referencia": "06/2026", "desoneracao": "SEM desoneração",
                 "observacao": "Custos diretos, BDI 0 — mesma natureza dos preços da DER-ES",
                 "arquivo": src.split('/')[-1]},
        "insumos": insumos, "composicoes": comps}
json.dump(base, open(out, 'w', encoding='utf-8'), ensure_ascii=False)
print(f"{out}: {len(insumos)} insumos + {len(comps)} composições com preço em {uf}")
grupos = {}
for c in comps:
    grupos[c['grp']] = grupos.get(c['grp'], 0) + 1
print(f"grupos de composição: {len(grupos)}")
