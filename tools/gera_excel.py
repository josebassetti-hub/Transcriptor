#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Gera a planilha Excel completa do orçamento a partir do orcamento.json do motor.

Uso:  python3 tools/gera_excel.py <orcamento.json> [saida.xlsx] [--refs data]
"""
import json, os, sys
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

argv = [a for a in sys.argv[1:] if not a.startswith('--')]
refs = 'data'
if '--refs' in sys.argv:
    refs = sys.argv[sys.argv.index('--refs') + 1]
src = argv[0] if argv else 'orcamento.json'
out = argv[1] if len(argv) > 1 else os.path.splitext(src)[0] + '.xlsx'

oj = json.load(open(src, encoding='utf-8'))
base = json.load(open(os.path.join(refs, 'base-der-es.json'), encoding='utf-8'))
regras = json.load(open(os.path.join(refs, 'regras-medicao.json'), encoding='utf-8'))['regras']
CAPS = base['capitulos']

def br(v, casas=2):
    """Formata número no padrão brasileiro (1.885,79)."""
    return f"{v:,.{casas}f}".replace(',', 'X').replace('.', ',').replace('X', '.')

TEAL = "0F766E"; TEAL_L = "D6F0EC"; CINZA = "F1F5F9"; LARANJA = "F97316"
RS = 'R$ #,##0.00'; NUM = '#,##0.00'; PCT = '0.0%'
thin = Side(style='thin', color='D0D7DE')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def head(ws, row, cols, fill=TEAL, color="FFFFFF"):
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = Font(bold=True, color=color, size=10)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 28

def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()
bdi = oj['totais']['bdi_pct'] / 100
CD = oj['totais']['custo_direto']; TOT_DER = oj['totais']['total']
COMPL = oj['complemento_a_cotar']; TOT_GERAL = oj['totais']['total_geral_com_complemento']
AREA = oj['obra']['area']

# ─────────────── 1. RESUMO ───────────────
ws = wb.active; ws.title = "1. Resumo"
widths(ws, [4, 52, 20, 18, 12])
ws['B2'] = "ORÇAMENTO DE OBRA"; ws['B2'].font = Font(bold=True, size=18, color=TEAL)
ws['B3'] = oj['obra']['nome']; ws['B3'].font = Font(bold=True, size=12)
ws['B4'] = oj['obra']['local']; ws['B4'].font = Font(size=10, color="475569")
r = 6
for k, v in [("Área construída", f"{br(AREA)} m²"),
             ("Pavimentos", oj['obra']['pav']),
             ("Padrão de acabamento", oj['obra']['padrao']),
             ("Base de preços", f"Tabela DER-ES · data-base {oj['obra'].get('data_base','')}"),
             ("BDI adotado", f"{br(oj['totais']['bdi_pct'], 1)}%"),
             ("Nº de itens orçados (DER-ES)", len(oj['itens'])),
             ("Nº de itens do complemento", len(COMPL['itens'])),
             ("  · com preço de outra tabela", sum(1 for i in COMPL['itens']
                                                   if 'cota' not in str(i.get('fonte', 'cotação')).lower())),
             ("  · por cotação de mercado", sum(1 for i in COMPL['itens']
                                                if 'cota' in str(i.get('fonte', 'cotação')).lower()))]:
    ws.cell(row=r, column=2, value=k).font = Font(bold=True, size=10)
    ws.cell(row=r, column=3, value=v).font = Font(size=10); r += 1

r += 1
ws.cell(row=r, column=2, value="RESUMO FINANCEIRO").font = Font(bold=True, size=12, color=TEAL); r += 1
head(ws, r, ["", "Descrição", "Valor (R$)", "% do total", ""]); r += 1
pbdi = br(oj['totais']['bdi_pct'], 1)
linhas = [("Custo direto — itens da tabela DER-ES", CD, False),
          (f"BDI {pbdi}% sobre o custo direto", CD * bdi, False),
          ("SUBTOTAL DER-ES (com BDI)", TOT_DER, True),
          ("Complemento — itens a cotar (custo direto)", COMPL['custo_direto'], False),
          (f"BDI {pbdi}% sobre o complemento", COMPL['custo_direto'] * bdi, False),
          ("SUBTOTAL COMPLEMENTO (com BDI)", COMPL['com_bdi'], True)]
for d, v, b in linhas:
    ws.cell(row=r, column=2, value=d).font = Font(bold=b, size=10)
    c = ws.cell(row=r, column=3, value=v); c.number_format = RS; c.font = Font(bold=b, size=10)
    p = ws.cell(row=r, column=4, value=v / TOT_GERAL); p.number_format = PCT; p.font = Font(bold=b, size=10)
    if b:
        for col in range(2, 5): ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=TEAL_L)
    for col in range(2, 5): ws.cell(row=r, column=col).border = BORDER
    r += 1
ws.cell(row=r, column=2, value="TOTAL GERAL DA OBRA").font = Font(bold=True, size=12, color="FFFFFF")
c = ws.cell(row=r, column=3, value=TOT_GERAL); c.number_format = RS; c.font = Font(bold=True, size=12, color="FFFFFF")
ws.cell(row=r, column=4, value=1).number_format = PCT
ws.cell(row=r, column=4).font = Font(bold=True, size=12, color="FFFFFF")
for col in range(2, 5): ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=TEAL)
r += 2
ws.cell(row=r, column=2, value="CUSTO POR METRO QUADRADO").font = Font(bold=True, size=11)
c = ws.cell(row=r, column=3, value=TOT_GERAL / AREA); c.number_format = RS; c.font = Font(bold=True, size=11, color=LARANJA)

# resumo por capítulo (+ gráfico de barras)
r += 3
ws.cell(row=r, column=2, value="DISTRIBUIÇÃO POR CAPÍTULO").font = Font(bold=True, size=12, color=TEAL); r += 1
head(ws, r, ["", "Capítulo", "Total c/ BDI (R$)", "% do total", ""]); r += 1
cap_tot = {}
for it in oj['itens']:
    cap_tot[it['c'][:2]] = cap_tot.get(it['c'][:2], 0) + it['total'] * (1 + bdi)
cap_tot['— COTAR'] = COMPL['com_bdi']
ini = r
for k in sorted(cap_tot, key=lambda x: -cap_tot[x]):
    nome = "Complemento a cotar (mercado)" if k == '— COTAR' else f"{k} · {CAPS.get(k, '')}"
    ws.cell(row=r, column=2, value=nome).font = Font(size=9)
    c = ws.cell(row=r, column=3, value=cap_tot[k]); c.number_format = RS; c.font = Font(size=9)
    c = ws.cell(row=r, column=4, value=cap_tot[k] / TOT_GERAL); c.number_format = PCT; c.font = Font(size=9)
    for col in range(2, 5): ws.cell(row=r, column=col).border = BORDER
    r += 1
ch = BarChart(); ch.type = "bar"; ch.style = 10
ch.title = "Participação por capítulo (R$ com BDI)"
ch.y_axis.title = None; ch.x_axis.title = None; ch.legend = None
ch.add_data(Reference(ws, min_col=3, min_row=ini, max_row=r - 1), titles_from_data=False)
ch.set_categories(Reference(ws, min_col=2, min_row=ini, max_row=r - 1))
ch.height = 12; ch.width = 18
ws.add_chart(ch, "F6")

r += 2
ws.cell(row=r, column=2, value="AVISO: estudo indicativo por metodologia paramétrica sobre a Tabela Referencial DER-ES "
        "e cotações de mercado. Não substitui orçamento executivo com projetos complementares, "
        "nem dispensa responsável técnico.").font = Font(italic=True, size=9, color="92400E")
ws.merge_cells(start_row=r, start_column=2, end_row=r + 2, end_column=4)
ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical='top')

# ─────────────── 2. PLANILHA ORÇAMENTÁRIA ───────────────
ws = wb.create_sheet("2. Planilha DER-ES")
widths(ws, [10, 62, 7, 12, 14, 14, 16])
ws['A1'] = "PLANILHA ORÇAMENTÁRIA — TABELA DER-ES"; ws['A1'].font = Font(bold=True, size=13, color=TEAL)
head(ws, 3, ["Código", "Especificação do serviço", "Und", "Quant.", "PU s/ BDI", "PU c/ BDI", "Total c/ BDI"])
ws.freeze_panes = "A4"
por_cap = {}
for it in oj['itens']: por_cap.setdefault(it['c'][:2], []).append(it)
r = 4
for cap in sorted(por_cap):
    itens = sorted(por_cap[cap], key=lambda x: x['c'])
    sub = sum(i['total'] for i in itens) * (1 + bdi)
    ws.cell(row=r, column=1, value=cap).font = Font(bold=True, color="FFFFFF", size=10)
    ws.cell(row=r, column=2, value=CAPS.get(cap, '')).font = Font(bold=True, color="FFFFFF", size=10)
    c = ws.cell(row=r, column=7, value=sub); c.number_format = RS; c.font = Font(bold=True, color="FFFFFF", size=10)
    for col in range(1, 8):
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=TEAL)
        ws.cell(row=r, column=col).border = BORDER
    r += 1
    for it in itens:
        vals = [it['c'], it['d'], it['u'], it['qtd'], it['pu'], it['pu'] * (1 + bdi), it['total'] * (1 + bdi)]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col, value=v); c.border = BORDER; c.font = Font(size=9)
            if col == 4: c.number_format = NUM
            if col in (5, 6, 7): c.number_format = RS
            if col == 2: c.alignment = Alignment(wrap_text=True, vertical='top')
        r += 1
ws.cell(row=r, column=2, value="TOTAL — ITENS DER-ES (com BDI)").font = Font(bold=True, size=11)
c = ws.cell(row=r, column=7, value=TOT_DER); c.number_format = RS; c.font = Font(bold=True, size=11)
for col in range(1, 8): ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=TEAL_L)

# ─────────────── 3. COMPLEMENTO A COTAR ───────────────
ws = wb.create_sheet("3. Complemento a cotar")
widths(ws, [50, 10, 6, 14, 16, 16, 15, 15, 26, 52])
ws['A1'] = "COMPLEMENTO — ITENS SEM PREÇO NA TABELA DER-ES"; ws['A1'].font = Font(bold=True, size=13, color=LARANJA)
ws['A2'] = COMPL['nota']; ws['A2'].font = Font(italic=True, size=9, color="475569")
ws.merge_cells('A2:J2'); ws.row_dimensions[2].height = 30
ws['A2'].alignment = Alignment(wrap_text=True, vertical='top')
head(ws, 4, ["Item", "Quant.", "Und", "Preço unit. (R$)", "Total s/ BDI", "Total c/ BDI",
             "Com preço de tabela (R$/un)", "A cotar (R$/un)", "Fonte do preço", "Observação"], fill=LARANJA)
ws.freeze_panes = "A5"
r = 5
for it in sorted(COMPL['itens'], key=lambda x: -x['total']):
    fonte = it.get('fonte', 'Cotação de mercado')
    mo = it.get('mo_oficial', 0)
    vals = [it['descricao'], it['qtd'], it['und'], it['pu'], it['total'], it['total'] * (1 + bdi),
            mo or None, (it.get('material_cotar') or None) if mo else it['pu'], fonte, it.get('obs', '')]
    for col, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=col, value=v); c.border = BORDER; c.font = Font(size=9)
        if col == 2: c.number_format = NUM
        if col in (4, 5, 6, 7, 8): c.number_format = RS
        if col in (1, 10): c.alignment = Alignment(wrap_text=True, vertical='top')
        if col == 7 and mo:
            c.fill = PatternFill("solid", fgColor="BBF7D0"); c.font = Font(bold=True, size=9)
        if col == 9 and 'SINAPI' in str(fonte):
            c.fill = PatternFill("solid", fgColor="BBF7D0"); c.font = Font(bold=True, size=9)
    r += 1
ws.cell(row=r, column=1, value="SUBTOTAL COMPLEMENTO").font = Font(bold=True, size=11)
for col, v in [(5, COMPL['custo_direto']), (6, COMPL['com_bdi'])]:
    c = ws.cell(row=r, column=col, value=v); c.number_format = RS; c.font = Font(bold=True, size=11)
for col in range(1, 11): ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="FED7AA")
if COMPL.get('mo_oficial_total'):
    ws.cell(row=r + 1, column=1, value="dos quais JÁ TÊM preço oficial de tabela (composição completa ou mão de obra):").font = Font(bold=True, size=10)
    c = ws.cell(row=r + 1, column=5, value=COMPL['mo_oficial_total']); c.number_format = RS
    c.font = Font(bold=True, size=10, color="15803D")
r += 2
ws.cell(row=r, column=1, value="RECOMENDAÇÃO: solicitar cotação de 3 fornecedores para os 5 maiores itens desta aba — "
        "concentram a maior parte da incerteza do orçamento.").font = Font(italic=True, size=9, color="92400E")

# ─────────────── 4. CURVA ABC ───────────────
ws = wb.create_sheet("4. Curva ABC")
widths(ws, [10, 62, 16, 12, 12, 8])
ws['A1'] = "CURVA ABC — TODOS OS ITENS (DER-ES + complemento)"; ws['A1'].font = Font(bold=True, size=13, color=TEAL)
head(ws, 3, ["Código", "Item", "Total c/ BDI", "% do total", "% acum.", "Classe"])
ws.freeze_panes = "A4"
todos = [{"c": i['c'], "d": i['d'], "t": i['total'] * (1 + bdi)} for i in oj['itens']]
todos += [{"c": "COTAR", "d": i['descricao'], "t": i['total'] * (1 + bdi)} for i in COMPL['itens']]
todos.sort(key=lambda x: -x['t'])
r = 4; ac = 0
for it in todos:
    p = it['t'] / TOT_GERAL; ac += p
    classe = "A" if ac <= 0.80 else ("B" if ac <= 0.95 else "C")
    cor = {"A": "FEE2E2", "B": "FEF3C7", "C": "ECFDF5"}[classe]
    for col, v in enumerate([it['c'], it['d'], it['t'], p, ac, classe], start=1):
        c = ws.cell(row=r, column=col, value=v); c.border = BORDER; c.font = Font(size=9)
        if col == 3: c.number_format = RS
        if col in (4, 5): c.number_format = PCT
        if col == 6:
            c.fill = PatternFill("solid", fgColor=cor); c.alignment = Alignment(horizontal='center')
            c.font = Font(bold=True, size=9)
        if col == 2: c.alignment = Alignment(wrap_text=True, vertical='top')
    r += 1

# ─────────────── 5. MEMORIAL DE CÁLCULO ───────────────
ws = wb.create_sheet("5. Memorial")
widths(ws, [10, 46, 12, 7, 46, 50])
ws['A1'] = "MEMORIAL DE CÁLCULO — origem de cada quantidade"; ws['A1'].font = Font(bold=True, size=13, color=TEAL)
head(ws, 3, ["Código", "Serviço", "Quant.", "Und", "Origem da quantidade (fórmula)", "Critério de medição — Caderno Técnico DER-ES"])
ws.freeze_panes = "A4"
r = 4
for it in oj['itens']:
    crit = (regras.get(it['c']) or {}).get('criterio') or "— caderno técnico deste item não ingerido —"
    for col, v in enumerate([it['c'], it['d'], it['qtd'], it['u'], it.get('formula', ''), crit], start=1):
        c = ws.cell(row=r, column=col, value=v); c.border = BORDER; c.font = Font(size=9)
        if col == 3: c.number_format = NUM
        if col in (2, 5, 6): c.alignment = Alignment(wrap_text=True, vertical='top')
    r += 1

# ─────────────── 6. PREMISSAS ───────────────
ws = wb.create_sheet("6. Premissas e lacunas")
widths(ws, [4, 110])
ws['B1'] = "PREMISSAS ADOTADAS"; ws['B1'].font = Font(bold=True, size=13, color=TEAL)
r = 3
for p in oj['premissas']:
    ws.cell(row=r, column=2, value="• " + p).alignment = Alignment(wrap_text=True, vertical='top')
    ws.cell(row=r, column=2).font = Font(size=10)
    ws.row_dimensions[r].height = max(15, 13 * (len(p) // 105 + 1)); r += 1
r += 2
ws.cell(row=r, column=2, value="LACUNAS E PONTOS DE ATENÇÃO").font = Font(bold=True, size=13, color="B45309"); r += 2
for l in oj['lacunas']:
    ws.cell(row=r, column=2, value="⚠ " + l).alignment = Alignment(wrap_text=True, vertical='top')
    ws.cell(row=r, column=2).font = Font(size=10)
    ws.row_dimensions[r].height = max(15, 13 * (len(l) // 105 + 1)); r += 1
r += 2
ws.cell(row=r, column=2, value="NÃO INCLUÍDOS NO ORÇAMENTO").font = Font(bold=True, size=13, color="B45309"); r += 2
for n in ["SPDA (sistema de proteção contra descargas atmosféricas) — requer projeto específico",
          "PPCI / instalações de combate a incêndio — requer projeto aprovado no Corpo de Bombeiros",
          "Elevador ou plataforma elevatória — NBR 9050 exige rota acessível vertical em edificação de "
          "uso público com 2 pavimentos; as pranchas mostram apenas escada (estimativa R$ 90–160 mil)",
          "WC PNE no 1º pavimento — previsto apenas no térreo",
          "Equipamentos de academia (esteiras, estações, catracas, TVs)",
          "Paisagismo, pavimentação do estacionamento e urbanização externa",
          "Mobiliário solto e decoração"]:
    ws.cell(row=r, column=2, value="• " + n).alignment = Alignment(wrap_text=True, vertical='top')
    ws.cell(row=r, column=2).font = Font(size=10)
    ws.row_dimensions[r].height = max(15, 13 * (len(n) // 105 + 1)); r += 1

# ─────────────── 7. AMBIENTES (take-off) ───────────────
ws = wb.create_sheet("7. Ambientes")
widths(ws, [30, 16, 12, 12, 14, 16])
ws['A1'] = "LEVANTAMENTO DE AMBIENTES (base do take-off)"; ws['A1'].font = Font(bold=True, size=13, color=TEAL)
head(ws, 3, ["Ambiente", "Tipo", "Área (m²)", "Perím. (m)", "Dist. quadro (m)", "Dist. prumada/cx (m)"])
r = 4; soma = 0
for a in oj['ambientes']:
    soma += a['area']
    for col, v in enumerate([a['nome'], a['tipo'], a['area'], a['per'], a.get('distQ'), a.get('distP')], start=1):
        c = ws.cell(row=r, column=col, value=v); c.border = BORDER; c.font = Font(size=9)
        if col >= 3: c.number_format = NUM
    r += 1
ws.cell(row=r, column=1, value="TOTAL ÁREA ÚTIL").font = Font(bold=True)
c = ws.cell(row=r, column=3, value=soma); c.number_format = NUM; c.font = Font(bold=True)
r += 2
ws.cell(row=r, column=1, value="Área construída total (c/ paredes e circulações):").font = Font(bold=True, size=10)
c = ws.cell(row=r, column=3, value=AREA); c.number_format = NUM; c.font = Font(bold=True, size=10)

# ─────────────── 8. REFERÊNCIAS CRUZADAS (opcional) ───────────────
XREF = oj.get('referencias_cruzadas')
if XREF:
    ws = wb.create_sheet("8. Referências cruzadas")
    widths(ws, [38, 9, 6, 13, 52, 52, 62])
    ws['A1'] = "BUSCA DOS ITENS DE COMPLEMENTO EM TABELAS REFERENCIAIS"
    ws['A1'].font = Font(bold=True, size=13, color=TEAL)
    ws['A2'] = XREF['nota']; ws['A2'].font = Font(italic=True, size=9, color="475569")
    ws.merge_cells('A2:G2'); ws.row_dimensions[2].height = 42
    ws['A2'].alignment = Alignment(wrap_text=True, vertical='top')
    head(ws, 4, ["Item do complemento", "Quant.", "Und", "Resolvido por"]
              + XREF.get('colunas', ["DER-ES", "SINAPI"]) + ["Ação adotada"])
    ws.freeze_panes = "A5"
    COR = {"DER-ES": "BBF7D0", "SINAPI": "BBF7D0", "MO SINAPI": "D9F99D", "PARCIAL": "FEF3C7",
           "ÂNCORA": "DBEAFE", "ALTERNATIVA": "E9D5FF", "NÃO EXISTE": "FEE2E2"}
    ordem = {"DER-ES": 0, "SINAPI": 1, "MO SINAPI": 2, "ALTERNATIVA": 3, "PARCIAL": 4,
             "ÂNCORA": 5, "NÃO EXISTE": 6}
    r = 5
    for it in sorted(XREF['itens'], key=lambda x: (ordem.get(x['situacao'], 9), -x['qtd'])):
        vals = [it['item'], it['qtd'], it['und'], it['situacao'],
                it.get('der', ''), it.get('sinapi', ''), it['acao']]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col, value=v); c.border = BORDER; c.font = Font(size=9)
            if col == 2: c.number_format = NUM
            if col in (1, 5, 6, 7): c.alignment = Alignment(wrap_text=True, vertical='top')
            if col == 4:
                c.fill = PatternFill("solid", fgColor=COR.get(it['situacao'], "FFFFFF"))
                c.font = Font(bold=True, size=9)
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[r] = ws.row_dimensions[r]
        ws.row_dimensions[r].height = 62
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="LEGENDA — DER-ES / SINAPI: o item saiu da cotação e passou a ter preço "
            "oficial da tabela indicada · MO SINAPI: a composição existe e a mão de obra tem preço oficial, "
            "mas o material não é pesquisado pelo IBGE — só ele vai a cotação · ALTERNATIVA: existe solução equivalente em tabela, mas com "
            "especificação diferente da projetada · PARCIAL: a tabela cobre parte do escopo · ÂNCORA: sem "
            "equivalente, mas há item próximo que valida a ordem de grandeza · NÃO EXISTE: ausente das duas "
            "tabelas, nem como insumo.").font = Font(italic=True, size=9, color="92400E")
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=7)
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical='top')

wb.save(out)
print("gerado:", out)
print("abas:", ", ".join(wb.sheetnames))
print(f"total geral: R$ {br(TOT_GERAL)}  |  R$/m²: {br(TOT_GERAL / AREA)}")
