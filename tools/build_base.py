#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_base.py — converte a Tabela Referencial DER-ES (XLSX oficiais) em JSON para o app.

Entrada  (fontes/):  tab_DER-EDIFICACOES_<AAAA>_<MM>_servicos.xlsx
                     tab_DER-EDIFICACOES_<AAAA>_<MM>_insumos.xlsx
                     tab_DER-EDIFICACOES_<AAAA>_<MM>_composicoes.xlsx
Saída    (data/):    base-der-es.json         (capítulos + serviços — embutido no app)
                     composicoes-resumo.json  (CPU resumida por serviço — uso analítico)
                     insumos.json             (insumos por categoria — uso analítico)

Uso:  python3 tools/build_base.py [--fontes DIR] [--out DIR]
Rotina anual: baixar o zip da tabela nova no site do DER-ES, substituir os XLSX em fontes/,
rodar este script e re-embutir data/base-der-es.json no orcamentista.html (bloco BASE_DER).
"""
import argparse, glob, json, os, re, sys, unicodedata

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl não instalado — rode: pip install openpyxl")


def norm(s):
    return " ".join(str(s).split()) if s is not None else ""


def find_file(fontes, kind):
    pats = glob.glob(os.path.join(fontes, f"*{kind}*.xlsx"))
    if not pats:
        sys.exit(f"arquivo de {kind} não encontrado em {fontes}")
    return pats[0]


def parse_servicos(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    meta = {"fonte": "DER-ES / LABOR-CT-UFES", "arquivo": os.path.basename(path)}
    capitulos, servicos = {}, []
    header_seen = False
    for row in ws.iter_rows(values_only=True):
        cells = [norm(c) for c in row]
        joined = " ".join(cells)
        if not header_seen:
            m = re.search(r"Data Base:\s*([A-Za-zç]+/\d{4})", joined)
            if m:
                meta["data_base"] = m.group(1)
            m = re.search(r"Leis Sociais:\s*([\d.,]+)", joined)
            if m:
                meta["leis_sociais_pct"] = float(m.group(1).replace(",", "."))
            m = re.search(r"BDI:\s*([\d.,]+)", joined)
            if m:
                meta["bdi_pct"] = float(m.group(1).replace(",", "."))
            if cells[0] == "Item":
                header_seen = True
            continue
        item, fonte, desc, und, _q, pu = (row + (None,) * 7)[:6]
        if item is None:
            continue
        code = str(item).lstrip("'").strip()
        if not re.fullmatch(r"\d{2,8}", code):
            continue
        if fonte in (None, ""):  # linha de grupo (capítulo/subcapítulo)
            capitulos[code] = norm(desc)
        else:
            try:
                preco = round(float(pu), 2)
            except (TypeError, ValueError):
                continue
            servicos.append({
                "c": code,                      # código DER
                "d": norm(desc),                # descrição
                "u": norm(und),                 # unidade
                "p": preco,                     # preço unitário (custo direto, BDI 0)
                "cap": code[:2], "sub": code[:4],
            })
    wb.close()
    return meta, capitulos, servicos


def parse_insumos(path):
    # layout: col0='Código (prefixo apóstrofo), col1=Descrição, col2=Und, col3=Preço;
    # linhas "Categoria: X" demarcam Mão-de-obra / Materiais / Equipamentos
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    insumos, cat = [], None
    for row in ws.iter_rows(values_only=True):
        c0 = norm(row[0])
        m = re.search(r"Categoria:\s*(.+)", c0)
        if m:
            cat = m.group(1)
            continue
        code = c0.lstrip("'")
        if not (cat and re.fullmatch(r"\d{6}", code)):
            continue
        try:
            preco = round(float(row[3]), 2)
        except (TypeError, ValueError, IndexError):
            continue
        insumos.append({"c": code, "d": norm(row[1]), "u": norm(row[2]), "p": preco, "cat": cat})
    wb.close()
    return insumos


SEC_MAP = {"Equipamento": "equip", "Mão-de-obra": "mo", "Material": "mat", "Serviços": "serv"}


def parse_composicoes(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    comps, cur, sec = {}, None, None
    for row in ws.iter_rows(values_only=True):
        c0 = norm(row[0])
        joined_row = [norm(c) for c in row]
        joined = " ".join(joined_row)
        m = re.match(r"Serviço:", c0)
        if m:
            m2 = re.search(r"(\d{6,8})\s*-\s*(.+)", joined)
            und = joined_row[-1] if joined_row[-1] else ""
            if m2:
                cur = {"c": m2.group(1), "d": m2.group(2).replace("Unidade: " + und, "").strip(),
                       "u": und, "equip": [], "mo": [], "mat": [], "serv": []}
                comps[cur["c"]] = cur
                sec = None
            continue
        if cur is None:
            continue
        if c0 in SEC_MAP and all(x == "" for x in joined_row[1:3]):
            sec = SEC_MAP[c0]
            continue
        if c0.startswith("Itens de incidência") or c0.startswith("Itens de transporte"):
            sec = None
            continue
        if c0.startswith("Custo Direto Total"):
            nums = [x for x in joined_row if re.fullmatch(r"-?[\d.,]+", x)]
            if nums:
                cur["custo_direto"] = float(nums[-1].replace(",", "."))
            sec = None
            continue
        if c0.startswith("(D) Produção"):
            nums = [x for x in joined_row if re.fullmatch(r"-?[\d.,]+", x)]
            if nums:
                cur["producao"] = float(nums[-1].replace(",", "."))
            continue
        if c0.startswith("(") or c0.startswith("Custo Horário") or c0.startswith("Preço Unitário") \
           or c0.startswith("BDI") or c0.startswith("Base:") or c0.startswith("Planilha"):
            if c0.startswith("(A)") or c0.startswith("(B)") or c0.startswith("(F)") or c0.startswith("(G)"):
                sec = None
            continue
        if sec and c0:
            cells = joined_row
            code = next((x for x in cells[1:] if re.fullmatch(r"\d{6}", x)), None)
            if not code:
                continue
            nums = [x for x in cells if re.fullmatch(r"-?\d[\d.,]*", x) and x != code]
            try:
                consumo = float(cells[8].replace(",", ".")) if cells[8] else (float(nums[-2].replace(",", ".")) if len(nums) >= 2 else None)
            except (ValueError, IndexError):
                consumo = None
            try:
                custo = float(nums[-1].replace(",", ".")) if nums else None
            except ValueError:
                custo = None
            idx = cells[1:].index(code) + 1
            und = cells[idx + 1] if idx + 1 < len(cells) else ""
            cur[sec].append({"c": code, "d": c0, "u": und, "q": consumo, "v": custo})
    wb.close()
    return comps


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--fontes", default="fontes")
    ap.add_argument("--out", default="data")
    args = ap.parse_args()
    os.makedirs(args.out, exist_ok=True)

    meta, capitulos, servicos = parse_servicos(find_file(args.fontes, "servicos"))
    base = {"meta": meta, "capitulos": capitulos, "servicos": servicos}
    with open(os.path.join(args.out, "base-der-es.json"), "w", encoding="utf-8") as f:
        json.dump(base, f, ensure_ascii=False, separators=(",", ":"))

    insumos = parse_insumos(find_file(args.fontes, "insumos"))
    with open(os.path.join(args.out, "insumos.json"), "w", encoding="utf-8") as f:
        json.dump(insumos, f, ensure_ascii=False, separators=(",", ":"))

    comps = parse_composicoes(find_file(args.fontes, "composicoes"))
    with open(os.path.join(args.out, "composicoes-resumo.json"), "w", encoding="utf-8") as f:
        json.dump(comps, f, ensure_ascii=False, separators=(",", ":"))

    # ---- validação ----
    print(f"meta: {meta}")
    print(f"capítulos/grupos: {len(capitulos)} | serviços: {len(servicos)} | insumos: {len(insumos)} | composições: {len(comps)}")
    idx = {s["c"]: s for s in servicos}
    golden = {"140701": 112.49, "151801": 238.03, "141409": 21.23, "010201": 16.92}
    ok = True
    for c, v in golden.items():
        got = idx.get(c, {}).get("p")
        status = "OK" if got == v else "FALHOU"
        if got != v:
            ok = False
        print(f"  golden {c}: esperado {v} obtido {got} [{status}]")
    caps = {}
    for s in servicos:
        caps[s["cap"]] = caps.get(s["cap"], 0) + 1
    print("serviços por capítulo:", dict(sorted(caps.items())))
    sizes = {f: os.path.getsize(os.path.join(args.out, f)) for f in
             ("base-der-es.json", "insumos.json", "composicoes-resumo.json")}
    print("tamanhos:", {k: f"{v/1024:.0f}KB" for k, v in sizes.items()})
    if not ok:
        sys.exit("VALIDAÇÃO FALHOU")
    print("validação OK")


if __name__ == "__main__":
    main()
